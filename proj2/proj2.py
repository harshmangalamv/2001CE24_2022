import streamlit as st
from io import BytesIO
import openpyxl
from openpyxl.styles import borders
from openpyxl.styles.borders import Border
from openpyxl.styles import PatternFill
from pyxlsb import open_workbook as open_xlsb
import pandas as pd
import os
import numpy as np
from datetime import datetime

st.title('Python Project')
st.subheader("Hello!")
start_time = datetime.now()
from openpyxl.styles.borders import Border, Side
def set_border(ws, cell_range):
    rows = ws[cell_range]
    side = Side(border_style='thin', color="FF000000")
    rows = list(rows)  # we convert iterator to list for simplicity, but it's not memory efficient solution
    max_y = len(rows) - 1  # index of the last row
    for pos_y, cells in enumerate(rows):
        max_x = len(cells) - 1  # index of the last cell
        for pos_x, cell in enumerate(cells):
            border = Border(
                left=cell.border.left,
                right=cell.border.right,
                top=cell.border.top,
                bottom=cell.border.bottom
            )
            if pos_x == 0:
                border.left = side
            if pos_x == max_x:
                border.right = side
            if pos_y == 0:
                border.top = side
            if pos_y == max_y:
                border.bottom = side

            # set new border only if it's one of the edge cells
            if pos_x == 0 or pos_x == max_x or pos_y == 0 or pos_y == max_y:
                cell.border = border
m=st.markdown("""
<style>
div.stButton > button:first-child {
    background-color: rgb(50, 168, 168);
    width: 25%;
    border-radius: 50px;
    display: block;
  margin-left: auto;
  margin-right: auto;
}
div.stDownloadButton > download_button:first-child {
    background-color: rgb(50, 86, 168);
    width: 25%;
    border-radius: 50px;
    display: block;
  margin-left: auto;
  margin-right: auto;
}

</style>""", unsafe_allow_html=True)
files = st.file_uploader('Choose a xlsx file',type=['xlsx'],accept_multiple_files=True)
mx=st.number_input("Enter mod value",max_value=7000)
button_compute = st.button("Calculate")

for file in files:
    #if(os.path.isfile(os.path.join(my_file_path,file))):
        # f=open(os.path.join(my_file_path,file),'r')
        #print(file)
    if file is not None and mx>0:
        filename=file.name
        filename=filename.split("xlsx")[0]
        try:
            df= pd.read_excel(file,nrows=200)
        except: 
            print("ERROR: COULD NOT FIND THE INPUT FILE!")
        mod=mx
        print("Your input file is as follows:")
        st.dataframe(df)
        pd.options.mode.chained_assignment = None  # default='warn'
        df.at[0,'U-Avg']=df['U'].mean()
        df.at[0,'V-Avg']=df['V'].mean()
        df.at[0,'W-Avg']=df['W'].mean()
        Uavg=df['U'].mean()
        Vavg=df['V'].mean()
        Wavg=df['W'].mean()
        df['U1']=df['U']-Uavg
        df['V1']=df['V']-Vavg
        df['W1']=df['W']-Wavg
        df[' ']=' '
        n=len(df)
        def findoct(x,y,z):
                    if x >= 0 and y >= 0 and z >= 0:
                        return 1 #octant 1
                    elif x < 0 and y >= 0 and z >= 0:
                        return 2 #octant 2
                    elif x < 0 and y < 0 and z >= 0:
                        return 3 #octant 3
                    elif x >= 0 and y < 0 and z >= 0:
                        return 4 #octant 4
                    elif x >= 0 and y >= 0 and z < 0:
                        return -1 #octant -1
                    elif x < 0 and y >= 0 and z < 0:
                        return -2 #octant -2
                    elif x < 0 and y < 0 and z < 0:
                        return -3 #octant -3
                    elif x >= 0 and y < 0 and z < 0:
                        return -4 #octant -4
        for  i in range(n):
            x = df['U1'][i]#x value
            y=  df['V1'][i]#y value
            z=  df['W1'][i]#z value
            df.at[i,"octant"]=int(findoct(x,y,z))#passing values in the function
            df['id']=' '#making new column id
            df['id'][2]='User Input'
            df['OctantID']=' '#new column OctantID is created
        #  df['OctantID'][0]='Overall Count'
            df['1']=0
            df['-1']=0
            df['2']=0
            df['-2']=0
            df['3']=0
            df['-3']=0
            df['4']=0
            df['-4']=0
            #we can find the frequencies of '1' '2' '3' '4'
            #by simply iterating from  0 to mod
            #we can follow similar approach for finding frequencies in further ranges
        mydict={
        df['1'][0]:0,
        df['2'][0]:0,
        df['-1'][0]:0,
        df['-2'][0]:0,
        df['3'][0]:0,
        df['-3'][0]:0,
        df['4'][0]:0,
        df['-4'][0]:0
        }
        n=len(df.index)
        for  i in range(n):
            x = df['U1'][i]
            y=  df['V1'][i]
            z=  df['W1'][i]
            if findoct(x,y,z)==1:
                df['1'][0]+=1
            if findoct(x,y,z)==-1:
                df['-1'][0]+=1
            if findoct(x,y,z)==2:
                df['2'][0]+=1
            if findoct(x,y,z)==-2:
                df['-2'][0]+=1
            if findoct(x,y,z)==3:
                df['3'][0]+=1
            if findoct(x,y,z)==-3:
                df['-3'][0]+=1
            if findoct(x,y,z)==4:
                df['4'][0]+=1
            if findoct(x,y,z)==-4:
                df['-4'][0]+=1
        i=0
        f=2
        prefixsum=0#for calclulating presum as we iterate through the column for total
        while i<n:
            ranged=min(mod,n-prefixsum)
            #here if we have not reached the final row it means we have to take range = mod
            # but as soon as we reach the last row the range is not mod but it is n-prefixsum(i,e : n-sum till now)
            mydict1={1:0,2:0,-1:0,-2:0,3:0,-3:0,4:0,-4:0}
            for j in range(ranged):
                mydict1[df['octant'][j+prefixsum]]+=1
            i+=mod
            # df['1'][f]=mydict
            df['1'][f]=mydict1[1]#storing value corresponding to octant 1 in the required row and column
            df['-1'][f]=mydict1[-1]
            df['2'][f]=mydict1[2]
            df['-2'][f]=mydict1[-2]
            df['3'][f]=mydict1[3]
            df['-3'][f]=mydict1[-3]
            df['4'][f]=mydict1[4]
            df['-4'][f]=mydict1[-4]
            f+=1
            prefixsum+=mod
            df['Octant 1']=' '
            df['Octant -1']=' '
            df['Octant 2']=' '
            df['Octant -2']=' '
            df['Octant 3']=' '
            df['Octant -3']=' '
            df['Octant 4']=' '
            df['Octant -4']=' '
            octant_name_id_mapping = {"1":"Internal outward interaction", "-1":"External outward interaction", "2":"External Ejection", "-2":"Internal Ejection", "3":"External inward interaction", "-3":"Internal inward interaction", "4":"Internal sweep", "-4":"External sweep"}
        df['OctantID'][1]="Mod:"+str(mod) #to print mod value as a string in OctantID 1st row
        import math
        p=n/mod
        p1=math.ceil(p) #print the total value of no of range rows which will be dataframe size/mod value
        m1={0:1,1:-1,2:2,3:-2,4:3,5:-3,6:4,7:-4}#dictionary for mapping index to octant value
        df['Rank1 Octant ID']=' '#making Rank1 Octant ID column
        df['Rank1 Octant Name']=' '#making Rank1 Octant Name column
        df['Octant.']=' '#naming new column as Octant.
        df['Octant.'][0]='1'#filling the octant column
        df['Octant.'][1]='-1'#same as above
        df['Octant.'][2]='2'
        df['Octant.'][3]='-2'
        df['Octant.'][4]='3'
        df['Octant.'][5]='-3'
        df['Octant.'][6]='4'
        df['Octant.'][7]='-4'
        df['Count of Rank 1 Mod values']=' '#making new column for storing count of rank 1 values
        for k in range(p1+2):
            vecp={}#for storing values as pair in form of key value pair
            if(k==1):#as 2nd column has to be blank in xlsx sheet
                continue
            for i in range(8):
                if(i==0):
                    vecp[df['1'][k]]=m1[i]#here we pairing the mod values with the corresponding octant value
                elif(i==1):
                    vecp[df['-1'][k]]=m1[i]
                elif(i==2):
                    vecp[df['2'][k]]=m1[i]
                elif(i==3):
                    vecp[df['-2'][k]]=m1[i]
                elif(i==4):
                    vecp[df['3'][k]]=m1[i]
                elif(i==5):
                    vecp[df['-3'][k]]=m1[i]
                elif(i==6):
                    vecp[df['4'][k]]=m1[i]
                elif(i==7):
                    vecp[df['-4'][k]]=m1[i]
            vecp=sorted(vecp.items())#sorting the pair values
            vecp.reverse()#we are reversing the sorted pairs such that we get the pairs in decreasing order of key value
            #this is done so as to get the index in the decreasing order of the key values
            rank=[item[1] for item in vecp]#for the position value
            #print(rank)
            for i in range(len(rank)):
                df.at[k, f"Octant {rank[i]}"] = i+1#pushing the position into the required dataframe column
                df.at[k, "Rank1 Octant ID"] = rank[0]#pushing the rank1 octant id which is rank[0], i.e the value of first key
                #which implies the octantID of rank1
                df.at[k, "Rank1 Octant Name"] = octant_name_id_mapping[f"{rank[0]}"]
            ro1=0#for storing count of rank1 for octant '1'
            rneg1=0#for storing the count of rank1 for octant'-1
            r2=0#so on...
            rneg2=0
            r3=0
            rneg3=0
            r4=0
            rneg4=0
            for i in range(2, p1+2):
                x=df['Rank1 Octant ID'][i];
                if(x==1):#if octant id of rank1 is 1
                    ro1+=1#incrementing the required counter
                elif(x==-1):#similar as above
                    rneg1+=1#incrementing the counter
                elif(x==2):
                    r2+=1
                elif(x==-2):
                    rneg2+=1
                elif(x==3):
                    r3+=1
                elif(x==-3):
                    rneg3+=1
                elif(x==4):
                    r4+=1
                elif(x==-4):
                    rneg4+=1
            for i in range(8):
                #we will push the value of counters into their respective position
                #into the 'Count of Rank1 Mod Values' column
                if(i==0):
                    df['Count of Rank 1 Mod values'][i]=ro1
                elif (i==1):
                    df['Count of Rank 1 Mod values'][i]=rneg1
                elif (i==2):
                    df['Count of Rank 1 Mod values'][i]=r2
                elif(i==3):
                    df['Count of Rank 1 Mod values'][i]=rneg2
                elif(i==4):
                    df['Count of Rank 1 Mod values'][i]=r3
                elif(i==5):
                    df['Count of Rank 1 Mod values'][i]=rneg3
                elif(i==6):
                    df['Count of Rank 1 Mod values'][i]=r4
                elif(i==7):
                    df['Count of Rank 1 Mod values'][i]=rneg4
        df['Octant Name']=' '#making new Column named 'Octant Name'
        df['Octant Name'][0]='Internal outward interation'#storing the respective values
        df['Octant Name'][1]='External outward interaction'#similarily for others
        df['Octant Name'][2]='External Ejection'
        df['Octant Name'][3]='Internal Ejection'
        df['Octant Name'][4]='External inward interaction'
        df['Octant Name'][5]='Internal inward interaction'
        df['Octant Name'][6]='Internal sweep'
        df['Octant Name'][7]='External sweep'
        df['OctantID'][1]="Mod:"+str(mod) #to print mod value as a string in OctantID 1st row
        import math
        p=n/mod
        p1=math.ceil(p) #print the total value of no of range rows which will be dataframe size/mod value
        prev=0
        ct=0
        f=2 #iterator for traversing
        i=mod-1
        for f in range(2,p1+2,1):
            if f==p1+1:
                df['OctantID'][f]=str(ct)+"-"+str(n-1) #for making initial range values
            else:
                df['OctantID'][f]=str(ct)+"-"+str(i) #if we reach last range then we need to print range from ct till the end row value
            ct=i #stores previous value of i
            i+=mod#i gets updated my the user input value of mod
            f+=1
        df['OctantID'][14]='Count'
        df['id'][15]='From'
        df['OctantID'][12]='Overall Transistion Count'
        df['1'][13]='To'
        df['1'][14]='1'
        df['-1'][14]='-1'
        df['2'][14]='2'
        df['-2'][14]='-2'
        df['3'][14]='3'
        df['-3'][14]='-3'
        df['4'][14]='4'
        df['-4'][14]='-4'
        arr=[[0 for i in range(8)] for j in range(8)]
        g=df['octant'].size
        for i in range(g-1):
            lt = df['octant'][i]
            rt = df['octant'][i+1]
            lc1 = 0
            rc1 = 0
            if(lt>0):
                lc1=2*lt-2
            else:
                lc1=2*abs(lt)-1
            if(rt>0):
                rc1=2*rt-2
            else:
                rc1=2*abs(rt)-1
            arr[int(lc1)][int(rc1)]+=1
        #print(arr)
        df['OctantID'][15]='1'
        df['OctantID'][16]='-1'
        df['OctantID'][17]='2'
        df['OctantID'][18]='-2'
        df['OctantID'][19]='3'
        df['OctantID'][20]='-3'
        df['OctantID'][21]='4'
        df['OctantID'][22]='-4'
        for i in range(15,15+8):
            for j in range(14,14+8):
                df.iat[i,j]=arr[i-15][j-14]
        df['OctantID'][26]='Mod transition count'
        df['1'][13]='To'
        df['1'][28]='1'
        df['-1'][28]='-1'
        df['2'][28]='2'
        df['-2'][28]='-2'
        df['3'][28]='3'
        df['-3'][28]='-3'
        df['4'][28]='4'
        df['-4'][28]='-4'
        df['OctantID'][28]='Count'
        df['OctantID'][29]='1'
        df['OctantID'][30]='-1'
        df['OctantID'][31]='2'
        df['OctantID'][32]='-2'
        df['OctantID'][33]='3'
        df['OctantID'][34]='-3'
        df['OctantID'][35]='4'
        df['OctantID'][36]='-4'
        df['OctantID'][26]='Mod transition count'
        df['1'][13]='To'
        df['1'][41]='1'
        df['-1'][41]='-1'
        df['2'][41]='2'
        df['-2'][41]='-2'
        df['3'][41]='3'
        df['-3'][41]='-3'
        df['4'][41]='4'
        df['-4'][41]='-4'
        df['OctantID'][41]='Count'
        df['OctantID'][42]='1'
        df['OctantID'][43]='-1'
        df['OctantID'][44]='2'
        df['OctantID'][45]='-2'
        df['OctantID'][46]='3'
        df['OctantID'][47]='-3'
        df['OctantID'][48]='4'
        df['OctantID'][49]='-4'
        df['OctantID'][39]='Mod transition count'
        df['OctantID'][52]='Mod transition count'
        df['OctantID'][65]='Mod transition count'
        df['OctantID'][78]='Mod transition count'
        df['OctantID'][91]='Mod transition count'
        df['id'][29]='From'
        df['id'][42]='From'
        df['id'][55]='From'
        df['id'][68]='From'
        df['id'][81]='From'
        df['id'][94]='From'
        df['1'][27]='To'
        df['1'][40]='To'
        df['1'][53]='To'
        df['1'][66]='To'
        df['1'][79]='To'
        df['1'][92]='To'
        df['1'][54]='1'
        df['-1'][54]='-1'
        df['2'][54]='2'
        df['-2'][54]='-2'
        df['3'][54]='3'
        df['-3'][54]='-3'
        df['4'][54]='4'
        df['-4'][54]='-4'
        df['OctantID'][54]='Count'
        df['OctantID'][55]='1'
        df['OctantID'][56]='-1'
        df['OctantID'][57]='2'
        df['OctantID'][58]='-2'
        df['OctantID'][59]='3'
        df['OctantID'][60]='-3'
        df['OctantID'][61]='4'
        df['OctantID'][62]='-4'
        df['1'][67]='1'
        df['-1'][67]='-1'
        df['2'][67]='2'
        df['-2'][67]='-2'
        df['3'][67]='3'
        df['-3'][67]='-3'
        df['4'][67]='4'
        df['-4'][67]='-4'
        df['OctantID'][67]='Count'
        df['OctantID'][68]='1'
        df['OctantID'][69]='-1'
        df['OctantID'][70]='2'
        df['OctantID'][71]='-2'
        df['OctantID'][72]='3'
        df['OctantID'][73]='-3'
        df['OctantID'][74]='4'
        df['OctantID'][75]='-4'
        df['1'][80]='1'
        df['-1'][80]='-1'
        df['2'][80]='2'
        df['-2'][80]='-2'
        df['3'][80]='3'
        df['-3'][80]='-3'
        df['4'][80]='4'
        df['-4'][80]='-4'
        df['OctantID'][80]='Count'
        df['OctantID'][81]='1'
        df['OctantID'][82]='-1'
        df['OctantID'][83]='2'
        df['OctantID'][84]='-2'
        df['OctantID'][85]='3'
        df['OctantID'][86]='-3'
        df['OctantID'][87]='4'
        df['OctantID'][88]='-4'
        df['1'][93]='1'
        df['-1'][93]='-1'
        df['2'][93]='2'
        df['-2'][93]='-2'
        df['3'][93]='3'
        df['-3'][93]='-3'
        df['4'][93]='4'
        df['-4'][93]='-4'
        df['OctantID'][93]='Count'
        df['OctantID'][94]='1'
        df['OctantID'][95]='-1'
        df['OctantID'][96]='2'
        df['OctantID'][97]='-2'
        df['OctantID'][98]='3'
        df['OctantID'][99]='-3'
        df['OctantID'][100]='4'
        df['OctantID'][101]='-4'
        prefixsum=0#for calclulating presum as we iterate through the column for total
        row=26
        n=df['U'].size
        while prefixsum<n:
            ranged=min(mod,n-prefixsum)
                #here if we have not reached the final row it means we have to take range = mod
                # but as soon as we reach the last row the range is not mod but it is n-prefixsum(i,e : n-sum till now)
            arr=[[0 for j in range(8)] for i in range(8)]
            for j in range(prefixsum,prefixsum+ranged-1):
                x=df.at[j,'octant']
                y=df.at[j+1,'octant']
                if x>0:
                    x=2*x-2
                else:
                    x=2*abs(x)-1
                if y>0:
                    y=2*y-2
                else:
                    y=2*abs(y)-1
                x=int(x)
                y=int(y)
                arr[x][y]+=1
            itr = row+3
        #     print(arr)
            for c in range(itr,itr+8):
                for j in range(14,14+8):
                    df.iat[c,j]=arr[c-itr][j-14]
            prefixsum+=mod
            row+= 13
        for  i in range(n):
            x = df['U1'][i]#x value
            y=  df['V1'][i]#y value
            z=  df['W1'][i]#z value
            df.at[i,"octant"]=int(findoct(x,y,z))#passing values in the function
        mydict={-1:0,1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0}#creating a dictionary to map values for count
        ct=0
        for j in range(-4,5):
            if(j==0): #0 is not allowed in column so we skip this case
                continue
            for i in range(n):
                if(df['octant'][i]==j):
                    ct+=1 #incrementing the count
                else:
                    mydict[j]=max(mydict[j],ct) #if a different value comes then we take max of ct and the dictionary mapped value if max is there dictionary is updated
                    ct=0  #counter if reinitialized to zero
        mydict2 ={-1:0,1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0} #creating another dictionary to map values for frequencies
        for j in range(-4,5):
            if(j==0): #skipping the loop when column value comes out to be zero (not allowed case)
                continue
            for i in range(n):
                if(df['octant'][i]==j):
                    ct+=1
                else:
                    if(ct==mydict[j]): #if the segment has count value equal to dictionary value it means we have to add this case to our answer
                        mydict2[j]+=1 #incrementing the frequency of such cases
                    ct=0
        df['Freq']=' '#creating frequency column
        df['count']=' '#creating count column
        df['count'][0]='+1'
        df['count'][1]='-1'
        df['count'][2]='+2'
        df['count'][3]='-2'
        df['count'][4]='+3'
        df['count'][5]='-3'
        df['count'][6]='+4'
        df['count'][7]='-4'
        df['Longest Subsequence Length']=' '#making longest subsequence column
        ct=0
        for i in range(1,5):
            if(i%2==1):#if i is odd
                df['Longest Subsequence Length'][ct]=mydict[i]
                df['Freq'][ct]=mydict2[i]
                ct+=1#(row counter incremented)
                df['Longest Subsequence Length'][ct]=mydict[i*-1]#pushing value for -i example is i=1 then here we are pushing value for -1
                df['Freq'][ct]=mydict2[i*-1]#same as above but pushing value for frequency column
                ct+=1#counter increment (row counter)
            else:
                df['Longest Subsequence Length'][ct]=mydict[i]
                df['Freq'][ct]=mydict2[i]
                ct+=1#counter increment
                df['Longest Subsequence Length'][ct]=mydict[i*-1]
                df['Freq'][ct]=mydict2[i*-1]
                ct+=1
        df['OCTANT']=' '#new column generation named OCTANT
        df['Longest Subsequence Length.']=' '#column generation named Longest Subsequence Length.
        df['Count.']=' '#new column named as Count.
        ans=[[],[],[],[],[],[],[],[]]#here we are using nested list for appending the list for their respective indexes
        listoct=[1,-1,2,-2,3,-3,4,-4]#storing the order of the octants in the list listoct(this list will later help us in placing the negative octant values as index cannot be negative!)
        for j in range(-4,5):
            if(j==0): #0 is not allowed in column so we skip this case
                continue
            for i in range(n):
                if(df['octant'][i]==j):
                    ct+=1 #incrementing the count
                else:
                    if(mydict[j]==ct):#if the counter is equal to th
                        if(j<0):
                            p=-1+(j*(-2))#as we see that the index of -1 is 1 of -2 is 3 ,so this general relation has been derived from here
                            ans[p].append(df['T'][i])#appending the final time at which we get the longest subseqeunce length
                        else:
                            if(j==1):
                                p=0 #if j==1 index in listoct of 1 is 0
                            if(j==2):
                                p=2 #index in listoct of 2 is 2
                            if(j==3):
                                p=4 #index in listoct of 3 is 4
                            if(j==4):
                                p=6 #index in listoct of 4 is 6
                            ans[p].append(df['T'][i])#appending the final time (for positive j values)
                    ct=0  #counter is reinitialized to zero
        r=0#variable r for row number
        #print(ans) we can print the ans (nested list) to check the final times for the repective octant and subsequence length
        for i in range(8):#loop from 0 to 7 for all octant values ,here we are using the listoct list to place the positive and negative octant values correctly
            df['OCTANT'][r]=listoct[i]
            df['Longest Subsequence Length.'][r]=mydict[listoct[i]]
            df['Count.'][r]=mydict2[listoct[i]]
            r+=1
            df['OCTANT'][r]='Time'
            df['Longest Subsequence Length.'][r]='From'
            df['Count.'][r]='To'
            r+=1
            sz=len(ans[i])#storing the size of the list(which are situated in the nested list ans)
            for j in range(sz):
                if(i==0):
                    df['Longest Subsequence Length.'][r]=ans[i][j]-(mydict[1]-1)*0.01-0.01#we can calculate the initial time by subtracting (0.01)times the subsequence length
                    #for example if final time of subsequence of some octant was y and the max length recorded was l so intiial TIME will be
                    #y-l+1 (will be the inital time)
                    # we are also subtracting 0.01 beacuse we are storing the final time value for the time at which a change occurs
                    # so the final time for which the last value at which the max length was stored will be final calculated- 0.01
                    df['Count.'][r]=ans[i][j]-0.01#this is the value stored in the list in ans list (the final time at which max length was recorded)
                elif(i==1):
                    df['Longest Subsequence Length.'][r]=ans[i][j]-(mydict[-1]-1)*0.01-0.01
                    df['Count.'][r]=ans[i][j]-0.01
                elif(i==2):
                    df['Longest Subsequence Length.'][r]=ans[i][j]-(mydict[2]-1)*0.01-0.01
                    df['Count.'][r]=ans[i][j]-0.01
                elif(i==3):
                    df['Longest Subsequence Length.'][r]=ans[i][j]-(mydict[-2]-1)*0.01-0.01
                    df['Count.'][r]=ans[i][j]-0.01
                elif(i==4):
                    df['Longest Subsequence Length.'][r]=ans[i][j]-(mydict[3]-1)*0.01-0.01
                    df['Count.'][r]=ans[i][j]-0.01
                elif(i==5):
                    df['Longest Subsequence Length.'][r]=ans[i][j]-(mydict[-3]-1)*0.01-0.01
                    df['Count.'][r]=ans[i][j]-0.01
                elif(i==6):
                    df['Longest Subsequence Length.'][r]=ans[i][j]-(mydict[4]-1)*0.01-0.01
                    df['Count.'][r]=ans[i][j]-0.01
                elif(i==7):
                    df['Longest Subsequence Length.'][r]=ans[i][j]-(mydict[-4]-1)*0.01-0.01
                    df['Count.'][r]=ans[i][j]-0.01
                j+=1#incrementing j so that we can move further into the specific list
                r+=1#row increment(line change)
        for  i in range(n):
            x = df['U1'][i]#x value
            y=  df['V1'][i]#y value
            z=  df['W1'][i]#z value
            df.at[i,"octant"]=int(findoct(x,y,z))#passing values in the function
        mydict={-1:0,1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0}#creating a dictionary to map values for count
        ct=0
        for j in range(-4,5):
            if(j==0): #0 is not allowed in column so we skip this case
                continue
            for i in range(n):
                if(df['octant'][i]==j):
                    ct+=1 #incrementing the count
                else:
                    mydict[j]=max(mydict[j],ct) #if a different value comes then we take max of ct and the dictionary mapped value if max is there dictionary is updated
                    ct=0  #counter if reinitialized to zero
        mydict2 ={-1:0,1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0} #creating another dictionary to map values for frequencies
        for j in range(-4,5):
            if(j==0): #skipping the loop when column value comes out to be zero (not allowed case)
                continue
            for i in range(n):
                if(df['octant'][i]==j):
                    ct+=1
                else:
                    if(ct==mydict[j]): #if the segment has count value equal to dictionary value it means we have to add this case to our answer
                        mydict2[j]+=1 #incrementing the frequency of such cases
                    ct=0
        df['Freq']=' '#creatinf frequency column
        df['count']=' '
        df['count'][0]='+1'
        df['count'][1]='-1'
        df['count'][2]='+2'
        df['count'][3]='-2'
        df['count'][4]='+3'
        df['count'][5]='-3'
        df['count'][6]='+4'
        df['count'][7]='-4'
        df['Longest Subsequence Length']=' '#making longest subsequence column
        ct=0
        for i in range(1,5):
            if(i%2==1):#if i is odd
                df['Longest Subsequence Length'][ct]=mydict[i]
                df['Freq'][ct]=mydict2[i]
                ct+=1#(row counter incremented)
                df['Longest Subsequence Length'][ct]=mydict[i*-1]#pushing value for -i example is i=1 then here we are pushing value for -1
                df['Freq'][ct]=mydict2[i*-1]#same as above but pushing value for frequency column
                ct+=1#counter increment (row counter)
            else:
                df['Longest Subsequence Length'][ct]=mydict[i]
                df['Freq'][ct]=mydict2[i]
                ct+=1
                df['Longest Subsequence Length'][ct]=mydict[i*-1]
                df['Freq'][ct]=mydict2[i*-1]
                ct+=1
        try:
            df.to_excel(f'{filename}octant_analysis_mod{mod}.xlsx',index=False)
        except:
            print("Output file could not be created")
        wb = openpyxl.load_workbook(f'{filename}octant_analysis_mod{mod}.xlsx')
        #wb = openpyxl.load_workbook('oh.xlsx')
        ws = wb.active
        from openpyxl.styles import PatternFill
        for rows in ws.iter_rows(min_row=0, max_row=9, min_col=23,max_col=33):
            x=0
            for cell in rows:
                if(cell.value==1):
                    yellow_pattern_fill = PatternFill(start_color='00FFFF00', end_color= '00FFFF00', fill_type="solid")
                    cell.fill=yellow_pattern_fill
                    border1 = borders.Side(style=None, color='FF000000', border_style='thin')
        border0 = borders.Side(style=None, color=None, border_style=None)
        thin = Border(left=border1, right=border0, bottom=border0, top=border0)
        set_border(ws, "N4:V9")
        set_border(ws, "N17:V24")
        set_border(ws, "N30:V38")
        set_border(ws, "N43:V51")
        set_border(ws, "N56:V64")
        set_border(ws, "N69:V77")
        set_border(ws, "N82:V90")
        set_border(ws, "N95:V103")
        set_border(ws,"AG2:AL9")
        set_border(ws,"AM2:AO27")
        set_border(ws,"W2:AE9")
        x=file
        wb.save(f'{filename}octant_analysis_mod{mod}.xlsx')
        with open(f'{filename}octant_analysis_mod{mod}.xlsx',"rb") as f:
            s=f.read()
        st.download_button(label='Download Current Result',data =s,file_name=f'{filename}octant_analysis_mod{mod}.xlsx')
end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))